from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SUMMARY_COLUMNS = [
    ("file_index", "Index"),
    ("filename", "Filename"),
    ("status", "Status"),
    ("annotation_path", "Annotation Path"),
    ("original_path", "Original Path"),
    ("point_count", "Point Count"),
    ("annotated_point_count", "Annotated Points"),
    ("annotated_percentage", "Annotated %"),
    ("unannotated_point_count", "Unannotated Points"),
    ("unannotated_percentage", "Unannotated %"),
    ("distinct_annotation_colors", "Distinct Colors"),
    ("is_dirty", "Dirty"),
    ("is_annotated", "Annotated (Saved)"),
    ("is_visited", "Visited"),
    ("is_copied_from_original", "Copied From Original"),
    ("brush_size_px", "Brush Size (px)"),
    ("point_size_px", "Point Size (px)"),
    ("annotation_alpha_pct", "Alpha %"),
    ("gamma_value", "Gamma"),
    ("render_points_as_spheres", "Points As Spheres"),
    ("file_format", "Format"),
    ("annotation_file_size_bytes", "File Size (bytes)"),
    ("annotation_file_mtime", "File Modified"),
    ("bbox_extent", "BBox Extent (x,y,z)"),
    ("centroid", "Centroid (x,y,z)"),
    ("comment", "Comment"),
    ("comment_updated_at", "Comment Updated"),
    ("stats_updated_at", "Stats Updated"),
]


def _format_cell(val):
    if isinstance(val, list):
        return ", ".join(f"{v:.3f}" if isinstance(v, float) else str(v) for v in val)
    return val


def _bold_header(ws) -> None:
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)


def export_review_to_excel(review_json_path: Path, dest_path: Path) -> int:
    """Build the review workbook (Summary / Class Breakdown / Session Info)
    from review.json. Returns the number of file entries exported."""
    data = json.loads(review_json_path.read_text(encoding="utf-8"))
    entries: dict = data.get("entries", {}) or {}
    meta: dict = data.get("meta", {}) or {}

    sorted_keys = sorted(entries.keys(), key=lambda k: entries[k].get("file_index", 0) or 0)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append([label for _, label in SUMMARY_COLUMNS])
    _bold_header(ws)
    for key in sorted_keys:
        e = entries[key]
        ws.append([_format_cell(e.get(field, "")) for field, _ in SUMMARY_COLUMNS])
    for i in range(1, len(SUMMARY_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws2 = wb.create_sheet("Class Breakdown")
    ws2.append(["Filename", "Class Name", "Color (Hex)", "Point Count", "Percentage"])
    _bold_header(ws2)
    for key in sorted_keys:
        e = entries[key]
        for cls in e.get("class_breakdown") or []:
            ws2.append([
                e.get("filename", ""),
                cls.get("name", ""),
                cls.get("hex", ""),
                cls.get("point_count", 0),
                cls.get("percentage", 0.0),
            ])
    for i in range(1, 6):
        ws2.column_dimensions[get_column_letter(i)].width = 20

    ws3 = wb.create_sheet("Session Info")
    ws3.append(["Field", "Value"])
    _bold_header(ws3)
    for k, v in meta.items():
        ws3.append([k, v])
    ws3.append(["Total Files Recorded", len(entries)])
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 40

    dest_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest_path))
    return len(entries)
